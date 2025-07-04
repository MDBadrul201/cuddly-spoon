#PDF Extracter
import openpyxl
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
import pandas as pd
import fitz
import threading
from queue import Queue
from concurrent.futures import ThreadPoolExecutor


print("Please wait while the UI is loading...")
print("Created by Mohd Badrul Amin")

# Function to extract specific pages from a PDF file using PyMuPDF
def extract_pages_with_mpn(pdf_file, mpn, output_dir, naming_convention, page_numbers=None):
    pdf_document = fitz.open(pdf_file)
    new_pdf = fitz.open()

    if page_numbers is None:
        page_numbers = []

    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        page_text = page.get_text("text")

        if re.search(str(mpn), page_text) or (page_numbers and (page_num + 1) in page_numbers):
            new_pdf.insert_pdf(pdf_document, from_page=page_num, to_page=page_num)

    if new_pdf.page_count > 0:
        output_file = os.path.join(output_dir, f"{naming_convention}.pdf")
        new_pdf.save(output_file)

    new_pdf.close()
    pdf_document.close()

# Function to browse for an excel file
def browse_for_file(entry_widget):
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, file_path)


# Function to browse for a pdf file
def browse_for_pdf(entry_widget):
    file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, file_path)


# Function to browse for a folder
def browse_for_folder(entry_widget):
    folder_path = filedialog.askdirectory()
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, folder_path)


def process_mpn(pdf_file, output_folder, naming_convention, mpn, page_numbers, progress_queue, result_queue):
    try:
        extract_pages_with_mpn(pdf_file, mpn, output_folder, naming_convention, page_numbers)
        progress_queue.put(1)
        if not page_numbers:
            result_queue.put(f"No pages found with MPN: {mpn} for {naming_convention}")
    except Exception as e:
        result_queue.put(f"Error processing {mpn} for {naming_convention}: {e}")


def extract_pages_worker(excel_file, pdf_file, output_folder, progress_queue, result_queue):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    total_rows = sheet.max_row - 1
    threads = []

    with ThreadPoolExecutor(max_workers=4) as executor:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 3:
                continue  # Skip rows that do not have at least three columns

            naming_convention, mpn, pages_str = row[:3]
            if isinstance(pages_str, int):
                # Convert integer to string and split
                pages_str = str(pages_str)
            elif not isinstance(pages_str, str):
                pages_str = ''  # Handle unexpected types

            if pages_str:
                try:
                    page_numbers = list(map(int, pages_str.split(',')))
                except ValueError:
                    page_numbers = []
            else:
                page_numbers = None

            future = executor.submit(process_mpn, pdf_file, output_folder, naming_convention, mpn, page_numbers, progress_queue, result_queue)
            threads.append(future)

    for t in threads:
        t.result()

    wb.close()
    result_queue.put("Extraction complete")



def update_progress_bar(total, progress_queue, result_queue):
    progress = 0
    while progress < total:
        progress += progress_queue.get()
        progress_bar['value'] = (progress / total) * 100
        root.update_idletasks()
    result_label.config(text=result_queue.get())


def extract_pages():
    excel_file = input_folder_entry.get()
    pdf_file = pdf_file_entry.get()
    output_folder = output_folder_entry.get()

    if not os.path.isfile(excel_file):
        result_label.config(text=f"Excel file not found: {excel_file}")
        return

    if not os.path.isfile(pdf_file):
        result_label.config(text=f"PDF file not found: {pdf_file}")
        return

    if not os.path.isdir(output_folder):
        result_label.config(text=f"Output folder does not exist: {output_folder}")
        return

    result_label.config(text="Processing...")

    progress_queue = Queue()
    result_queue = Queue()

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    total_rows = sheet.max_row - 1
    wb.close()

    extraction_thread = threading.Thread(target=extract_pages_worker, args=(excel_file, pdf_file, output_folder, progress_queue, result_queue))
    extraction_thread.start()

    progress_thread = threading.Thread(target=update_progress_bar, args=(total_rows, progress_queue, result_queue))
    progress_thread.start()


def generate_template():
    df = pd.DataFrame(columns=['NAMING CONVENTION', 'MPN', 'PAGES'])

    data = {'NAMING CONVENTION': ['PDF NAMING GOES HERE'],
            'MPN': ['MPN GOES HERE'],
            'PAGES': ['1,2,3']}  # Example of pages format

    df = pd.DataFrame(data)

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

    if file_path:
        df.to_excel(file_path, index=False)
        messagebox.showinfo("Template Created", "Template Excel file created successfully.")
        os.system(f'start excel "{file_path}"')

def close_app():
    root.destroy()

def launch_pdf_extractor():
    global root, input_folder_entry, pdf_file_entry, output_folder_entry, result_label, progress_bar
    root = tk.Tk()
    root.title("PDF Page Extractor")
    root.geometry("600x400")  # Set the window size to 400x400

    instructions_label = tk.Label(root, text="Click the 'Create Template' button to create a template Excel file.\nEdit the Excel file and click the 'Extract Pages' button to extract pages from the PDF file.\nEnsure that the PDF file and Excel file is closed before clicking the 'Extract Pages' button.\nCreated by Mohd Badrul Amin B. Safary")
    instructions_label.grid(row=0, column=0, columnspan=4, padx=20, pady=10, sticky='w')
    input_folder_entry = tk.Entry(root, width=50)
    input_folder_label = tk.Label(root, text="Input Excel File:")
    input_folder_label.grid(row=2, column=0, sticky='w', padx=20, pady=10)
    input_folder_entry.grid(row=2, column=1, padx=10, pady=10, columnspan=2, sticky='w')
    input_browse_button = tk.Button(root, text="Browse", command=lambda: browse_for_file(input_folder_entry))
    input_browse_button.grid(row=2, column=3, padx=10, pady=10, sticky='w')

    pdf_file_entry = tk.Entry(root, width=50)
    pdf_file_label = tk.Label(root, text="PDF File:")
    pdf_file_label.grid(row=3, column=0, sticky='w', padx=20, pady=10)
    pdf_file_entry.grid(row=3, column=1, padx=10, pady=10, columnspan=2, sticky='w')
    pdf_browse_button = tk.Button(root, text="Browse", command=lambda: browse_for_pdf(pdf_file_entry))
    pdf_browse_button.grid(row=3, column=3, padx=10, pady=10, sticky='w')

    output_folder_entry = tk.Entry(root, width=50)
    output_folder_label = tk.Label(root, text="Output Folder:")
    output_folder_label.grid(row=4, column=0, sticky='w', padx=20, pady=10)
    output_folder_entry.grid(row=4, column=1, padx=10, pady=10, columnspan=2, sticky='w')
    output_browse_button = tk.Button(root, text="Browse", command=lambda: browse_for_folder(output_folder_entry))
    output_browse_button.grid(row=4, column=3, padx=10, pady=10, sticky='w')

    create_template_button = tk.Button(root, text="Create Template", command=generate_template, bg="blue", fg="white")
    create_template_button.grid(row=1, column=0, columnspan=4, padx=20, pady=10, sticky='w')

    extract_button = tk.Button(root, text="Extract Pages", command=extract_pages, bg="green", fg="white")
    extract_button.grid(row=5, column=0, columnspan=4, padx=20, pady=10, sticky='w')

    result_label = tk.Label(root, text="", wraplength=400)
    result_label.grid(row=6, column=0, columnspan=4, padx=20, pady=10, sticky='w')

    progress_bar = Progressbar(root, orient=tk.HORIZONTAL, length=400, mode='determinate')
    progress_bar.grid(row=7, column=0, columnspan=4, padx=20, pady=10, sticky='w')

    close_app_button = tk.Button(root, text="Close", command=close_app, bg="red", fg="black")
    close_app_button.grid(row=8, column=0, columnspan=4, padx=20, pady=10, sticky='w')

    root.mainloop()

if __name__ == "__main__":
    launch_pdf_extractor()
